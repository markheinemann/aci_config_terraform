import csv
import yaml
from collections import defaultdict
import subprocess
import os
import shutil
import openpyxl
import pprint
import sys



##############################################
#          PROCESS tenants            #
##############################################

xlsx_file_path = "Book2.xlsx"
worksheet_name = "TENANT"

output_yaml_file_path_leaf = "tenant_vars.yml"

data_tenant = defaultdict(list)

# Open the Excel file
workbook = openpyxl.load_workbook(xlsx_file_path)

# Select the worksheet by name
worksheet = workbook[worksheet_name]

# Iterate through rows in the worksheet
for row in worksheet.iter_rows(min_row=2, values_only=True):
    type = row[0]  # Assuming "type" is in the first column
    print(type)

    if type == "tenant":
        print("match tenant")
        data_tenant[type].append({
            "tenant_name": row[1],  # Assuming "tenant_name" is in the second column
            "tenant_description": row[2]  # Assuming "tenant_description" is in the third column
        })


with open(output_yaml_file_path_leaf, 'w') as yaml_file:
    for function, items in data_tenant.items():
        yaml_file.write("{}:\n".format(function))
        for item in items:
#            print(item)
            yaml_file.write("- tenant: {}\n".format(item["tenant_name"]))



            yaml_file.write("  description: {}\n".format(item["tenant_description"]))

            yaml_file.write("\n")




# Close the Excel file
#workbook.close()

pprint.pprint(data_tenant)

# #sys.exit()

##############################################
#          PROCESS vrf           #
##############################################

xlsx_file_path = "Book2.xlsx"
worksheet_name = "VRF"

output_yaml_file_path_leaf = "vrf_vars.yml"

data_vrf = defaultdict(list)

# Open the Excel file
workbook = openpyxl.load_workbook(xlsx_file_path)

# Select the worksheet by name
worksheet = workbook[worksheet_name]

# Iterate through rows in the worksheet
for row in worksheet.iter_rows(min_row=2, values_only=True):
    type = row[0]  # Assuming "type" is in the first column
    print(type)

    if type == "vrf":
        
        data_vrf[type].append({
            "vrf_name": row[1],  
            "vrf_tenant": data_tenant["tenant"][0]["tenant_name"]
        })


with open(output_yaml_file_path_leaf, 'w') as yaml_file:
    for function, items in data_vrf.items():
        yaml_file.write("{}:\n".format(function))
        for item in items:
#            print(item)
            yaml_file.write("- vrf: {}\n".format(item["vrf_name"]))
            yaml_file.write("  vrf_tenant: {}\n".format(item["vrf_tenant"]))
            yaml_file.write("\n")




# Close the Excel file
#workbook.close()

pprint.pprint(data_vrf)



##############################################
#          PROCESS bd            #
##############################################

xlsx_file_path = "Book2.xlsx"
worksheet_name = "BD"

output_yaml_file_path_leaf = "bd_vars.yml"

data_bd = defaultdict(list)

# Open the Excel file
#workbook = openpyxl.load_workbook(xlsx_file_path)
worksheet = workbook[worksheet_name]

for row in worksheet.iter_rows(min_row=2, values_only=True):
    type = row[0]  # Assuming "type" is in the first column
    print(type)

    if type == "bd":
        #print("match")
        data_bd[type].append({
            "bd_tenant_name": data_tenant["tenant"][0]["tenant_name"],
            "bd_name": row[1],
            "bd_description": row[2],
            "bd_vrf": data_vrf['vrf'][0]['vrf_name']
        })


with open(output_yaml_file_path_leaf, 'w') as yaml_file:
    for function, items in data_bd.items():
        yaml_file.write("{}:\n".format(function))
        for item in items:
#            print(item)
            yaml_file.write("- bd: {}\n".format(item["bd_name"]))
            yaml_file.write("  bd_tenant_name: {}\n".format(item["bd_tenant_name"]))
            yaml_file.write("  bd_description: {}\n".format(item["bd_description"]))
            yaml_file.write("  bd_vrf: {}\n".format(item["bd_vrf"]))
            yaml_file.write("\n")

pprint.pprint(data_bd)






##############################################
#          PROCESS subnet          #
##############################################
xlsx_file_path = "Book2.xlsx"
worksheet_name = "SUBNET"

output_yaml_file_path_leaf = "subnet_vars.yml"

data_subnet = defaultdict(list)

# Open the Excel file
#workbook = openpyxl.load_workbook(xlsx_file_path)
worksheet = workbook[worksheet_name]

for row in worksheet.iter_rows(min_row=2, values_only=True):
    type = row[0]  # Assuming "type" is in the first column
    print(type)

    if type == "subnet":
        data_subnet[type].append({
            "subnet_name": row[1],
            "subnet_tn_name": data_tenant["tenant"][0]["tenant_name"],
            "subnet_bd": row[2],
            "subnet_scope": row[3],
            "subnet_description": row[4]
        })

with open(output_yaml_file_path_leaf, 'w') as yaml_file:
    for function, items in data_subnet.items():
        yaml_file.write("{}:\n".format(function))
        for item in items:
#            print(item)
            yaml_file.write("- subnet: {}\n".format(item["subnet_name"]))
            yaml_file.write("  subnet_tn_name: {}\n".format(item["subnet_tn_name"]))
            yaml_file.write("  subnet_bd: {}\n".format(item["subnet_bd"]))
            yaml_file.write("  subnet_scope: {}\n".format(item["subnet_scope"]))
            yaml_file.write("  subnet_description: {}\n".format(item["subnet_description"]))
            yaml_file.write("\n")


#
pprint.pprint(data_subnet)













###

##############################################
#          PROCESS application_profile        #
##############################################

xlsx_file_path = "Book2.xlsx"
worksheet_name = "APP_PROFILE"

output_yaml_file_path_leaf = "app_profile_vars.yml"

data_app_profile = defaultdict(list)

# Open the Excel file
#workbook = openpyxl.load_workbook(xlsx_file_path)
worksheet = workbook[worksheet_name]

for row in worksheet.iter_rows(min_row=2, values_only=True):
    type = row[0]  # Assuming "type" is in the first column
    print(type)


    if type == "app_profile":
            data_app_profile[type].append({
                "app_profile_name": row[1],
                "app_profile_tenant": data_tenant["tenant"][0]["tenant_name"]
            })

with open(output_yaml_file_path_leaf, 'w') as yaml_file:
    for function, items in data_app_profile.items():
        yaml_file.write("{}:\n".format(function))
        for item in items:
#            print(item)
            yaml_file.write("- app_profile: {}\n".format(item["app_profile_name"]))
            yaml_file.write("  app_profile_tenant: {}\n".format(item["app_profile_tenant"]))


            yaml_file.write("\n")



pprint.pprint(data_app_profile)






##############################################
#          PROCESS epg       #
##############################################

xlsx_file_path = "Book2.xlsx"
worksheet_name = "EPG"

output_yaml_file_path_leaf = "epg_vars.yml"

data_epg = defaultdict(list)

# Open the Excel file
#workbook = openpyxl.load_workbook(xlsx_file_path)
worksheet = workbook[worksheet_name]

for row in worksheet.iter_rows(min_row=2, values_only=True):
    type = row[0]  # Assuming "type" is in the first column
    print(type)



    if type == "epg":
        data_epg[type].append({
            "epg_name": row[1],
            "epg_tenant": data_tenant["tenant"][0]["tenant_name"],
            "epg_map_to_bd": row[2],
            "epg_map_to_app_profile": row[3]
        })

with open(output_yaml_file_path_leaf, 'w') as yaml_file:
    for function, items in data_epg.items():
        yaml_file.write("{}:\n".format(function))
        for item in items:
#            print(item)
            yaml_file.write("- epg: {}\n".format(item["epg_name"]))
            yaml_file.write("  epg_tenant: {}\n".format(item["epg_tenant"]))
            yaml_file.write("  epg_map_to_bd: {}\n".format(item["epg_map_to_bd"]))
            yaml_file.write("  epg_map_to_app_profile: {}\n".format(item["epg_map_to_app_profile"]))

            
            yaml_file.write("\n")

pprint.pprint(data_epg)




##############################################
#          PROCESS VLAN_Pool     #
##############################################

xlsx_file_path = "Book2.xlsx"
worksheet_name = "VLAN_POOL"

output_yaml_file_path_leaf = "vlan_pool_vars.yml"

data_vlan_pool = defaultdict(list)

# Open the Excel file
#workbook = openpyxl.load_workbook(xlsx_file_path)
worksheet = workbook[worksheet_name]

for row in worksheet.iter_rows(min_row=2, values_only=True):
    type = row[0]  # Assuming "type" is in the first column
    print(type)



    if type == "vlan_pool":
        data_vlan_pool[type].append({
            "vlan_pool_name": row[1],
            #"epg_tenant": data_tenant["tenant"][0]["tenant_name"],
            "range_from": row[2],
            "range_to": row[3],
            "allocation_mode": row[4]
        })

with open(output_yaml_file_path_leaf, 'w') as yaml_file:
    for function, items in data_vlan_pool.items():
        yaml_file.write("{}:\n".format(function))
        for item in items:
#            print(item)
            yaml_file.write("- vlan_pool: {}\n".format(item["vlan_pool_name"]))
            #yaml_file.write("  vlan_pool_name: {}\n".format(item["vlan_pool_name"]))
            yaml_file.write("  range_from: {}\n".format(item["range_from"]))
            yaml_file.write("  range_to: {}\n".format(item["range_to"]))
            yaml_file.write("  allocation_mode: {}\n".format(item["allocation_mode"]))
            
            yaml_file.write("\n")

pprint.pprint(data_vlan_pool)




##############################################
#          PROCESS PHYSICAL_DOMAIN    #
##############################################

xlsx_file_path = "Book2.xlsx"
worksheet_name = "PHYSICAL_DOMAIN"

output_yaml_file_path_leaf = "physical_domain_vars.yml"

data_physical_domain = defaultdict(list)

# Open the Excel file
#workbook = openpyxl.load_workbook(xlsx_file_path)
worksheet = workbook[worksheet_name]

for row in worksheet.iter_rows(min_row=2, values_only=True):
    type = row[0]  # Assuming "type" is in the first column
    print(type)



    if type == "physical_domain":
        data_physical_domain[type].append({
            "physical_domain_name": row[1],
            "mode": row[3],
            "vlan_pool_name": row[2]
        })

with open(output_yaml_file_path_leaf, 'w') as yaml_file:
    for function, items in data_physical_domain.items():
        yaml_file.write("{}:\n".format(function))
        for item in items:
#            print(item)
            yaml_file.write("- physical_domain: {}\n".format(item["physical_domain_name"]))
            #yaml_file.write("  vlan_pool_name: {}\n".format(item["vlan_pool_name"]))
            yaml_file.write("  vlan_pool_name: {}\n".format(item["vlan_pool_name"]))
            yaml_file.write("  mode: {}\n".format(item["mode"]))

            
            yaml_file.write("\n")

pprint.pprint(data_physical_domain)



#############################################
#          PROCESS AAEP   #
##############################################

xlsx_file_path = "Book2.xlsx"
worksheet_name = "AAEP"

output_yaml_file_path_leaf = "aaep_vars.yml"

data_aaep = defaultdict(list)

# Open the Excel file
#workbook = openpyxl.load_workbook(xlsx_file_path)
worksheet = workbook[worksheet_name]

for row in worksheet.iter_rows(min_row=2, values_only=True):
    type = row[0]  # Assuming "type" is in the first column
    print(type)



    if type == "aaep":
        data_aaep[type].append({
            "aaep_name": row[1],
            "attached_domain": row[2]
        })

with open(output_yaml_file_path_leaf, 'w') as yaml_file:
    for function, items in data_aaep.items():
        yaml_file.write("{}:\n".format(function))
        for item in items:
#            print(item)
            yaml_file.write("- aaep: {}\n".format(item["aaep_name"]))
            #yaml_file.write("  vlan_pool_name: {}\n".format(item["vlan_pool_name"]))
            yaml_file.write("  attached_domain: {}\n".format(item["attached_domain"]))
            
            yaml_file.write("\n")

pprint.pprint(data_aaep)





#############################################
#          PROCESS LINK LEVEL POLICY      #
##############################################

xlsx_file_path = "Book2.xlsx"
worksheet_name = "LINK_LEVEL_POLICY"

output_yaml_file_path_leaf = "link_level_policy_vars.yml"

data_link_level_policy = defaultdict(list)

# Open the Excel file
#workbook = openpyxl.load_workbook(xlsx_file_path)
worksheet = workbook[worksheet_name]

for row in worksheet.iter_rows(min_row=2, values_only=True):
    type = row[0]  # Assuming "type" is in the first column
    print(type)



    if type == "link_level_policy":
        data_link_level_policy[type].append({
            "name": row[1],
            "description": row[2],
            "speed": row[3],
            "linkdebounce": row[4],
            "fec": row[5],
            "autoneg": row[6]
        })

with open(output_yaml_file_path_leaf, 'w') as yaml_file:
    for function, items in data_link_level_policy.items():
        yaml_file.write("{}:\n".format(function))
        for item in items:
#            print(item)
            yaml_file.write("- link_level_policy: {}\n".format(item["name"]))
            #yaml_file.write("  vlan_pool_name: {}\n".format(item["vlan_pool_name"]))
            yaml_file.write("  description: {}\n".format(item["description"]))
            yaml_file.write("  speed: {}\n".format(item["speed"]))
            yaml_file.write("  linkdebounce: {}\n".format(item["linkdebounce"]))
            yaml_file.write("  fec: {}\n".format(item["fec"]))
            yaml_file.write("  autoneg: '{}'\n".format(item["autoneg"]))
            
            
            yaml_file.write("\n")

pprint.pprint(data_link_level_policy)





#############################################
#          PROCESS CDP INT POLICY      #
##############################################

xlsx_file_path = "Book2.xlsx"
worksheet_name = "CDP_POLICY"

output_yaml_file_path_leaf = "cdp_policy_vars.yml"

cdp_policy = defaultdict(list)

# Open the Excel file
#workbook = openpyxl.load_workbook(xlsx_file_path)
worksheet = workbook[worksheet_name]

for row in worksheet.iter_rows(min_row=2, values_only=True):
    type = row[0]  # Assuming "type" is in the first column
    print(type)



    if type == "cdp_policy":
        cdp_policy[type].append({
            "name": row[1],
            "description": row[2],
            "admin_state": row[3]

        })

with open(output_yaml_file_path_leaf, 'w') as yaml_file:
    for function, items in cdp_policy.items():
        yaml_file.write("{}:\n".format(function))
        for item in items:
#            print(item)
            yaml_file.write("- cdp_policy: {}\n".format(item["name"]))
            #yaml_file.write("  vlan_pool_name: {}\n".format(item["vlan_pool_name"]))
            yaml_file.write("  description: {}\n".format(item["description"]))
            yaml_file.write("  admin_state: {}\n".format(item["admin_state"]))

            
            
            yaml_file.write("\n")

pprint.pprint(cdp_policy)


#############################################
#          PROCESS LACP INT POLICY      #
##############################################

xlsx_file_path = "Book2.xlsx"
worksheet_name = "LACP_POLICY"

output_yaml_file_path_leaf = "lacp_policy_vars.yml"

lacp_policy = defaultdict(list)

# Open the Excel file
#workbook = openpyxl.load_workbook(xlsx_file_path)
worksheet = workbook[worksheet_name]

for row in worksheet.iter_rows(min_row=2, values_only=True):
    type = row[0]  # Assuming "type" is in the first column
    print(type)



    if type == "lacp_policy":
        lacp_policy[type].append({
            "name": row[1],
            "description": row[2],
            "min_links": row[3],
            "max_links": row[4],
            "mode": row[5]

        })

with open(output_yaml_file_path_leaf, 'w') as yaml_file:
    for function, items in lacp_policy.items():
        yaml_file.write("{}:\n".format(function))
        for item in items:
#            print(item)
            yaml_file.write("- lacp_policy: {}\n".format(item["name"]))
            #yaml_file.write("  vlan_pool_name: {}\n".format(item["vlan_pool_name"]))
            yaml_file.write("  description: {}\n".format(item["description"]))
            yaml_file.write("  min_links: {}\n".format(item["min_links"]))
            yaml_file.write("  max_links: {}\n".format(item["max_links"]))
            yaml_file.write("  mode: {}\n".format(item["mode"]))

            
            
            yaml_file.write("\n")

pprint.pprint(lacp_policy)



#############################################
#          PROCESS VPC                       #
##############################################

xlsx_file_path = "Book2.xlsx"
worksheet_name = "VPC"

output_yaml_file_path_leaf = "vpc_vars.yml"

vpc_config = defaultdict(list)

# Open the Excel file
#workbook = openpyxl.load_workbook(xlsx_file_path)
worksheet = workbook[worksheet_name]

for row in worksheet.iter_rows(min_row=2, values_only=True):
    type = row[0]  # Assuming "type" is in the first column
    print(type)

    if type == "vpc_policy_group":
        vpc_config[type].append({
            "vpc_policy_group_name": row[1],
            "associated_cdp_policy": row[2],
            "associated_lacp_policy": row[3],
            "associated_aaep": row[4]
        })

    

    if type == "leaf_interface_profile":
        vpc_config[type].append({
            "leaf_interface_profile_name": row[1],
            "leaf_interface_profile_desc": row[2]
        })


    if type == "interface_selector":
        vpc_config[type].append({
            "interface_selector_name": row[1],
            "from_port": row[2],
            "to_port": row[3]
        })
    if type == "leaf_name":
        vpc_config[type].append({
            "leaf_name": row[1],
            "from_switch": row[2],
            "to_switch": row[3]
        })
print("hello")
pprint.pprint(vpc_config)

#sys.exit()



# with open(output_yaml_file_path_leaf, 'w') as yaml_file:
#     for function, items in vpc_config.items():
#         yaml_file.write("{}:\n".format(function))
#         for item in items:
#             print("Debug - item:", item)

#             ucs_fi_value = item['vpc_policy_group_name']
#             yaml_file.write("  - vpc_policy_group: {}\n".format(ucs_fi_value))


            # yaml_file.write("- vpc_policy_group: {}\n".format(item["vpc_policy_group_name"]))
            # yaml_file.write("  associated_cdp_policy: {}\n".format(item["associated_cdp_policy"]))
            # yaml_file.write("  associated_lacp_policy: {}\n".format(item["associated_lacp_policy"]))
            # yaml_file.write("  associated_aaep: {}\n".format(item["associated_aaep"]))
           
            # yaml_file.write("- leaf_interface_profile_name: {}\n".format(item["leaf_interface_profile_name"]))
            # #yaml_file.write("  vlan_pool_name: {}\n".format(item["vlan_pool_name"]))
            # yaml_file.write("  leaf_interface_profile_desc: {}\n".format(item["leaf_interface_profile_desc"]))

            # yaml_file.write("- interface_selector: {}\n".format(item["interface_selector_name"]))
            # #yaml_file.write("  vlan_pool_name: {}\n".format(item["vlan_pool_name"]))
            # yaml_file.write("  from_port: {}\n".format(item["from_port"]))
            # yaml_file.write("  to_port: {}\n".format(item["to_port"]))

            # yaml_file.write("- leaf_name: {}\n".format(item["leaf_name"]))
            # #yaml_file.write("  vlan_pool_name: {}\n".format(item["vlan_pool_name"]))
            # yaml_file.write("  from_switch: {}\n".format(item["from_switch"]))
            # yaml_file.write("  to_switch: {}\n".format(item["to_switch"]))          





            
            
        #yaml_file.write("\n")

#pprint.pprint(vpc_config)











# # Close the Excel file
workbook.close()


